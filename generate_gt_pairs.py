#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import os
import sys
import re

from openpyxl import load_workbook
import random

import pandas as pd

try:
    from openai import OpenAI
except ImportError:
    OpenAI = None

try:
    from google import genai
except ImportError:
    genai = None

try:
    import anthropic
except ImportError:
    anthropic = None

import spot
spot.setup()

import config
import json

os.environ['OPENAI_API_KEY'] = config.OPENAI_API_KEY
os.environ["GEMINI_API_KEY"] = config.GEMINI_API_KEY
os.environ["ANTHROPIC_API_KEY"] = config.ANTHROPIC_API_KEY

from utils import collect_aps


PROMPT = """Translate the following natural language requirement into Linear Temporal Logic (LTL).

Use STRICT programming-style notation:
- F for "finally"
- G for "globally"
- X for "next"
- U for "until"
- ! for negation
- & for AND
- | for OR
- -> for implication
- <-> for equivalence (iff)

Do NOT use LaTeX or any mathematical symbols such as ◇, □, ¬, ∧, ∨, etc.
Do NOT use the equality sign.
Do NOT include explanations, comments, or multiple formulas.

Use the following atomic proposition mapping:
{atomic_proposition}

Return EXACTLY one LTL formula as a single line.
The output must be directly parseable as an LTL formula.

Requirement:
{requirement}
"""


def normalize_formula(text: str) -> str:
    """Strip common formatting so the result is parser-friendly."""
    text = text.strip()

    if text.startswith("```"):
        lines = text.splitlines()
        lines = [line for line in lines if not line.strip().startswith("```")]
        text = "\n".join(lines).strip()

    # Keep only the first non-empty line if the model accidentally adds extra text.
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return lines[0] if lines else ""



def ask_claude(
    client: str,
    requirement: str,
    atomic_proposition: str,
) -> str:

    prompt = PROMPT.format(
        requirement=requirement,
        atomic_proposition=atomic_proposition,
    )

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=128,
        temperature=0,
        messages=[
            {
                "role": "user",
                "content": prompt,
            }
        ],
    )

    text = response.content[0].text

    return normalize_formula(text)



def ask_gemini(
    client: genai.Client,
    requirement: str,
    atomic_proposition: str,
) -> str:

    prompt = PROMPT.format(
        requirement=requirement,
        atomic_proposition=atomic_proposition,
    )

    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=prompt,
    )

    return normalize_formula(response.text or "")


def ask_chatgpt(client: OpenAI, requirement: str, atomic_proposition: str) -> str:
    response = client.chat.completions.create(
        model="gpt-5.4-mini",
        temperature=0,
        messages=[
            {
                "role": "user",
                "content": PROMPT.format(
                    requirement=requirement,
                    atomic_proposition=atomic_proposition,
                ),
            }
        ],
    )

    return normalize_formula(response.choices[0].message.content or "")


def semantically_equivalent(formula_a: str, formula_b: str):
    """
    Returns:
        True  -> semantically equivalent
        False -> valid syntax but not equivalent
        None  -> syntax error, exclude from accuracy
    """
    try:
        f_a = spot.formula(formula_a)
        f_b = spot.formula(formula_b)

        xor_formula = spot.formula.Not(spot.formula.Equiv(f_a, f_b))
        return spot.translate(xor_formula).is_empty()

    except Exception as exc:
        msg = str(exc)

        if "syntax error" in msg.lower():
            print(
                f"Syntax error; excluding from accuracy:\n"
                f"  Error:        {exc}",
                file=sys.stderr,
            )
            return None

        print(
            f"Warning: could not compare formulas:\n"
            f"  Error:        {exc}",
            file=sys.stderr,
        )
        return None


def extract_ap_mapping(ltl_formula: str) -> str:
    """
    Extract AP names from the GT formula using Spot.
    Return them as a set of strings (stringified for the prompt).
    """
    f = spot.formula(ltl_formula)
    aps = sorted(str(ap) for ap in spot.atomic_prop_collect(f))

    return "{" + ", ".join(aps) + "}"



def load_dataset(input_path: str):
    """
    Parse an input dataset file (.txt / .xlsx / .csv / spacewire.json / .json)
    into a list of (requirement, ground_truth, atomic_proposition) tuples.

    Returns (dataset, parse_errors).
    """

    dataset = []
    parse_errors = 0



    if input_path.endswith("txt"):



        def normalize_ltl_variables(formula):
            """
            Aggressive normalization of variable names and operators.
            """

            # ------------------------------------------------------------
            # Normalize temporal operators
            # ------------------------------------------------------------
            formula = formula.replace("<>", "F ")
            formula = formula.replace("[]", "G ")

            formula = formula.replace("AG", "G ")
            formula = formula.replace("AF", "F ")

            formula = formula.replace("EG", "G ")
            formula = formula.replace("EF", "F ")

            # Convert:
            # A[exp1 U exp2]  -->  exp1 U exp2
            formula = re.sub(r"A\[(.*?)\]", r"\1", formula)
            formula = re.sub(r"E\[(.*?)\]", r"\1", formula)

            # ------------------------------------------------------------
            # Normalize boolean operators
            # ------------------------------------------------------------
            formula = formula.replace("||", "|")
            formula = formula.replace("&&", "&")

            def replace_function_var(match):
                func_name = match.group(1)
                args = match.group(2)

                # Split args by comma and normalize
                parts = [p.strip() for p in args.split(",")]

                joined = "_".join(parts)

                return f"{func_name}_{joined}"

            formula = re.sub(
                r"\b([A-Za-z_][A-Za-z0-9_]*)\(([^()]+)\)",
                replace_function_var,
                formula,
            )

            formula = formula.replace(".", "_")
            formula = formula.replace("=", "_")
            formula = formula.replace("+", "_")

            formula = formula.replace("[", "_")
            formula = formula.replace("]", "")
            formula = formula.replace(":", "_")

            # Collapse repeated underscores
            formula = re.sub(r"_+", "_", formula)


            # ------------------------------------------------------------
            # Repair malformed parentheses
            #
            # Removes unmatched closing parens.
            # Adds missing closing parens at end.
            # ------------------------------------------------------------
            repaired = []
            open_count = 0

            for ch in formula:

                if ch == "(":
                    open_count += 1
                    repaired.append(ch)

                elif ch == ")":
                    if open_count > 0:
                        open_count -= 1
                        repaired.append(ch)
                    else:
                        # Skip unmatched closing paren
                        continue
                else:
                    repaired.append(ch)

            # Add missing closing parentheses
            repaired.append(")" * open_count)

            formula = "".join(repaired)


            # Remove spaces around underscores
            formula = re.sub(r"\s*_\s*", "_", formula)

            return formula.strip()



        FIELD_NAMES = {
            "REQUIREMENT:",
            "REFINEMENT:",
            "PATTERN:",
            "SCOPE:",
            "PARAMETERS:",
            "LTL:",
            "CTL:",
            "NOTE:",
            "SOURCE:",
            "DOMAIN:",
            "ORIGINAL",
            "REWRITING",
            "NOTE"
        }

        def starts_with_field(line):
            return any(line.startswith(field) for field in FIELD_NAMES)




        with open(input_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Split entries by blank lines
        # (works for typical structured requirement datasets)
        raw_items = re.split(r"\n\s*\n", content)

        malformed_count = 0

        for idx, item in enumerate(raw_items, start=1):

            try:
                requirement = None
                refinement = None
                ltl = None

                lines = item.splitlines()

                i = 0

                while i < len(lines):

                    line = lines[i].strip()

                    # --------------------------------------------------------
                    # REQUIREMENT
                    # --------------------------------------------------------
                    if line.startswith("REQUIREMENT:"):
                        requirement = line[len("REQUIREMENT:"):].strip()

                    # --------------------------------------------------------
                    # REFINEMENT
                    # --------------------------------------------------------
                    elif line.startswith("REFINEMENT:"):
                        refinement = line[len("REFINEMENT:"):].strip()

                    # --------------------------------------------------------
                    # LTL (possibly multiline)
                    # --------------------------------------------------------
                    elif line.startswith("LTL:"):

                        ltl_lines = [
                            line[len("LTL:"):].strip()
                        ]

                        i += 1

                        while i < len(lines):

                            next_line = lines[i].strip()

                            # Stop if another field begins
                            if starts_with_field(next_line):
                                i -= 1
                                break

                            if next_line:
                                ltl_lines.append(next_line)

                            i += 1

                        ltl = " ".join(ltl_lines)

                    i += 1

                # ------------------------------------------------------------
                # Validate required fields
                # ------------------------------------------------------------
                if not requirement:
                    raise ValueError("Missing REQUIREMENT")

                if not ltl:
                    raise ValueError("Missing LTL")

                # ------------------------------------------------------------
                # Combine REQUIREMENT + REFINEMENT
                # ------------------------------------------------------------
                nl_text = requirement

                if refinement:
                    nl_text += f" ({refinement})"

                # ------------------------------------------------------------
                # Normalize LTL syntax
                # ------------------------------------------------------------
                ltl = normalize_ltl_variables(ltl)

                try:
                    ap_set = extract_ap_mapping(ltl)
                except Exception as exc:
                    print(
                        f"Skipping item {idx} due to Spot parse error:\n"
                        f"  Formula: {ltl}\n"
                        f"  Error:   {exc}",
                        file=sys.stderr,
                    )
                    parse_errors += 1
                    continue

                dataset.append((nl_text, ltl, ap_set))

            except Exception as e:
                malformed_count += 1
                print(f"Malformed item {idx}: {e}")


    elif input_path.endswith("xlsx"):


        wb = load_workbook(input_path, data_only=True)
        ws = wb.active

        # ------------------------------------------------------------
        # Detect NL column dynamically from header row
        # ------------------------------------------------------------
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        nl_col_idx = None

        for idx, value in enumerate(header_row):
            if value == "NL":
                nl_col_idx = idx
                break

        if nl_col_idx is None:
            raise ValueError('Could not find header column named "NL"')

        # Relative offsets from NL column
        OFFSET_F = 4   # decision3
        OFFSET_G = 5   # bool_expr3 / bool_expr4
        OFFSET_H = 6   # decision2
        OFFSET_I = 7   # bool_expr2
        OFFSET_J = 8   # decision1
        OFFSET_K = 9   # bool_expr1

        pairs = []
        malformed_count = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):

            try:
                nl_text = row[nl_col_idx].value  # Column B

                # ------------------------------------------------------------
                # STEP 1-3 : decision2 / bool_expr2
                # ------------------------------------------------------------
                ltl_prefix = None
                used_bool_expr2 = False
                bool_expr2 = None

                h_val = row[nl_col_idx + OFFSET_H].value  # Column H

                if h_val:
                    h_json = json.loads(h_val)
                    decision2 = h_json[0]["decision2"]

                    if "upon bool_exp2" in decision2[0]:
                        i_val = row[nl_col_idx + OFFSET_I].value  # Column I

                        if not i_val:
                            raise ValueError("Missing column I")

                        i_json = json.loads(i_val)

                        bool_expr2 = i_json[0]["bool_exp2"][0]
                        used_bool_expr2 = True

                # ------------------------------------------------------------
                # STEP 4-6 : decision1 / bool_expr1
                # ------------------------------------------------------------
                j_val = row[nl_col_idx + OFFSET_J].value  # Column J

                if not j_val:
                    raise ValueError("Missing column J")

                j_json = json.loads(j_val)
                decision1 = j_json[0]["decision1"]

                bool_expr1 = None

                two_brackets = False

                if "while bool_exp1" in decision1[0] or "whenever bool_exp1" in decision1[0]:
                    k_val = row[nl_col_idx + OFFSET_K].value  # Column K

                    if not k_val:
                        raise ValueError("Missing column K")

                    k_json = json.loads(k_val)
                    bool_expr1 = k_json[0]["bool_exp1"][0]

                    if used_bool_expr2:
                        ltl_prefix = (
                            f"!{bool_expr2} U ({bool_expr2} & G({bool_expr1} -> "
                        )
                        two_brackets = True
                    else:
                        ltl_prefix = f"G({bool_expr1} -> "

                else:

                    if "upon bool_exp1" in decision1[0]:

                        k_val = row[nl_col_idx + OFFSET_K].value  # Column K

                        if not k_val:
                            raise ValueError("Missing column K")

                        k_json = json.loads(k_val)
                        bool_expr1 = k_json[0]["bool_exp1"][0]

                        ltl_prefix = (
                            f"!{bool_expr1} U ({bool_expr1} & "
                        )

                    elif used_bool_expr2:
                        ltl_prefix = (
                            f"!{bool_expr2} U ({bool_expr2} & "
                        )
                    else:

                        ltl_prefix = "("

                        # raise ValueError("No usable decision1/decision2 logic")

                # ------------------------------------------------------------
                # STEP 7-8 : decision3 random selection
                # ------------------------------------------------------------
                f_val = row[nl_col_idx + OFFSET_F].value  # Column F

                if not f_val:
                    raise ValueError("Missing column F")

                f_json = json.loads(f_val)
                decision3 = f_json[0]["decision3"]

                valid_choices = [
                    s for s in decision3
                    if "N_DURATION" not in s
                ]

                if not valid_choices:
                    raise ValueError("No valid decision3 entries")

                chosen = random.choice(valid_choices)

                # ------------------------------------------------------------
                # STEP 9 : bool_expr3 / bool_expr4
                # ------------------------------------------------------------
                g_val = row[nl_col_idx + OFFSET_G].value  # Column G

                if not g_val:
                    raise ValueError("Missing column G")

                g_json = json.loads(g_val)
                g_first = g_json[0]

                bool_expr3 = g_first["bool_exp3"][0]

                # ------------------------------------------------------------
                # STEP 10-11 : finish formula
                # ------------------------------------------------------------
                suffix = None

                if chosen == "eventually satisfy bool_exp3":
                    suffix = f"F({bool_expr3})"

                elif chosen == "always satisfy bool_exp3":
                    suffix = f"G({bool_expr3})"

                elif chosen == "at the next timepoint satisfy bool_exp3":
                    suffix = f"X({bool_expr3})"

                elif chosen == "immediately satisfy bool_exp3":
                    suffix = f"({bool_expr3})"

                elif chosen == "until bool_exp4, satisfy bool_exp3":
                    bool_expr4 = g_first["bool_exp4"][0]
                    suffix = f"({bool_expr3} U {bool_expr4})"

                else:
                    raise ValueError(f"Unsupported decision3 option: {chosen}")

                ltl_formula = ltl_prefix + suffix + ")"
                if two_brackets:
                    ltl_formula += ")"

                try:
                    ap_set = extract_ap_mapping(ltl_formula)
                except Exception as exc:
                    print(
                        f"Skipping item {row_idx} due to Spot parse error:\n"
                        f"  Formula: {ltl_formula}\n"
                        f"  Error:   {exc}",
                        file=sys.stderr,
                    )
                    parse_errors += 1
                    break

                dataset.append((nl_text, ltl_formula, ap_set))

            except Exception as e:
                malformed_count += 1
                print(f"Malformed row {row_idx}: {e}")



    elif input_path.endswith("csv"):

        df = pd.read_csv(input_path, sep=',')

        required_columns = {"Natural Language", "Ground Truth", "Atomic Proposition"}
        missing = required_columns - set(df.columns)
        if missing:
            raise ValueError(f"Missing required columns: {sorted(missing)}")

        for _, row in df.iterrows():
            requirement = str(row["Natural Language"])
            ground_truth = str(row["Ground Truth"]).strip()
            atomic_proposition = str(row["Atomic Proposition"]).strip()

            dataset.append((requirement, ground_truth, atomic_proposition))


    elif "spacewire.json" in input_path:

        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, list):
            raise ValueError("Input JSON must contain an array of objects.")


        for idx, item in enumerate(data):
            requirement = str(item.get("text", "")).strip()

            logic_entries = item.get("logics", [])

            if not isinstance(logic_entries, list):
                continue

            found_ltl = False

            for logic in logic_entries:
                if logic.get("type") != "LTL":
                    continue

                found_ltl = True

                f_code = logic.get("f_code", "")

                if not f_code or not str(f_code).strip():
                    break

                ground_truth = str(f_code).strip()

                ground_truth = ground_truth.replace("-->", "->")
                ground_truth = re.sub(r"\bnot\b", "!", ground_truth)
                ground_truth = re.sub(r"\band\b", "&", ground_truth)
                ground_truth = re.sub(r"\bor\b", "&", ground_truth)

                try:
                    ap_set = extract_ap_mapping(ground_truth)
                except Exception as exc:
                    print(
                        f"Skipping item {idx} due to Spot parse error:\n"
                        f"  Formula: {ground_truth}\n"
                        f"  Error:   {exc}",
                        file=sys.stderr,
                    )
                    parse_errors += 1
                    break

                dataset.append((requirement, ground_truth, ap_set))

                # only take the first valid LTL formula
                break

            if not found_ltl:
                continue
                                  

    elif input_path.endswith("json"):

        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if not isinstance(data, list):
            raise ValueError("Input JSON must contain an array of objects.")


        for idx, item in enumerate(data):
            if "nlTask" not in item or "ltlequ" not in item:
                raise ValueError(f"Element {idx} is missing 'nlTask' or 'ltlequ'.")

            requirement = str(item["nlTask"]).strip()
            ground_truth = str(item["ltlequ"][0]).strip()[:-1]

            # dataset cleanup before Spot parsing
            ground_truth = ground_truth.replace("U", " U ")

            try:
                atomic_proposition = extract_ap_mapping(ground_truth)
            except Exception as exc:
                print(
                    f"Ground-truth parse error at item {idx}; excluding:\n"
                    f"  Ground Truth: {ground_truth}\n"
                    f"  Error:        {exc}",
                    file=sys.stderr,
                )
                parse_errors += 1
                continue

            dataset.append((requirement, ground_truth, atomic_proposition))

    return dataset, parse_errors


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Translate natural language requirements to LTL and compare with ground truth."
    )
    parser.add_argument("input", help="Input file")
    parser.add_argument("output_csv", help="Output CSV file")
    # parser.add_argument(
    #     "--model",
    #     default="gpt-5.4-mini",
    #     help="OpenAI model to use, default: gpt-5.4-mini",
    # )

    args = parser.parse_args()

    # if not os.getenv("OPENAI_API_KEY"):
    #     raise RuntimeError("OPENAI_API_KEY environment variable is not set.")
    if not os.getenv("ANTHROPIC_API_KEY"):
        raise RuntimeError("ANTHROPIC_API_KEY environment variable is not set.")

    dataset, parse_errors = load_dataset(args.input)

    # client = OpenAI()

    # client = genai.Client(
    #     api_key=os.environ["GEMINI_API_KEY"]
    # )

    client = anthropic.Anthropic(
        api_key=os.environ["ANTHROPIC_API_KEY"]
    )

    rows = []
    correct = 0
    total = 0
    syntax_errors = 0

    seen = set()


    for requirement, ground_truth, atomic_proposition in dataset:

        # model_response = ask_chatgpt(client, requirement, atomic_proposition)
        # model_response = ask_gemini(client, requirement, atomic_proposition)
        model_response = ask_claude(client, requirement, atomic_proposition)
        

        # Spot validation timeout
        if ground_truth == (
            "!(newPatient) U ((newPatient) & F(((patientAttributesEntered & SelfTestMode) & "
            "((((((testPowerSwitchPass & testLeaksPass) & testFl2Pass) & testPSExpPass) & "
            "testOxygenSensorPass) & testAlarmsPass) -> selfTestPass))))"
        ):
            continue

        equivalent = semantically_equivalent(ground_truth, model_response)

        print(
            f"  Requirement: {requirement}\n"
            f"  Ground Truth: {ground_truth}\n"
            f"  Response:     {model_response}\n"
            f"  Equivalent:     {equivalent}\n",
            file=sys.stderr,
        )

        if equivalent is None:
            syntax_errors += 1
        else:
            pair = (ground_truth, model_response)
            if pair not in seen:

                total += 1
                seen.add(pair)

                if equivalent == False:
                    rows.append(
                        {
                            "Ground Truth": ground_truth,
                            "Response": model_response,
                        }
                    )

                else:
                    correct += 1

    with open(args.output_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["Ground Truth", "Response"],
        )
        writer.writeheader()
        writer.writerows(rows)

    accuracy = correct / total if total else 0.0
    print(f"Total accuracy: {accuracy:.4f} ({correct}/{total})")
    print(f"Syntax errors excluded: {syntax_errors}")


if __name__ == "__main__":
    main()